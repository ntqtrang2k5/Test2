import openpyxl
import json
import sys

# Set encoding to utf-8 for stdout
sys.stdout.reconfigure(encoding='utf-8')

file = "01_TaiLieu/TC_HopDong.xlsx"
wb = openpyxl.load_workbook(file, data_only=True)
print(f"Sheets: {wb.sheetnames}")

all_data = {}
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_data = []
    # Read first 50 rows of each sheet to be sure
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        sheet_data.append([str(cell) if cell is not None else "" for cell in row])
    all_data[sheet_name] = sheet_data

with open("scratch/hopdong_sheets_inspect.json", "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print("Done")
