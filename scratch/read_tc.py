import openpyxl
import json

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    results = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        results[sheet_name] = data
    return results

file_path = r"d:\03_Projects\LTW\Test2\01_TaiLieu\TC_ChinhSuaThongTinXe.xlsx"
content = read_excel(file_path)
with open(r"d:\03_Projects\LTW\Test2\scratch\tc_content.json", "w", encoding="utf-8") as f:
    json.dump(content, f, ensure_ascii=False, indent=2)
print("Content saved to scratch/tc_content.json")
