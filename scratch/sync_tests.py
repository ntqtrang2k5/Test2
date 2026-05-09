import openpyxl
import re
import os
import sys

# Set encoding to utf-8 for stdout
sys.stdout.reconfigure(encoding='utf-8')

# Mapping of Excel files and Sheets to test files
CARS_MAPPING = {
    "01_TaiLieu/TC_ThemThongTinXe.xlsx": {
        "prefix": "C", "file": "tests/playwright/cars/test_add.py"
    },
    "01_TaiLieu/TC_ChinhSuaThongTinXe.xlsx": {
        "prefix": "U", "file": "tests/playwright/cars/test_edit.py"
    },
    "01_TaiLieu/TC_XemVaLocThongTinXe.xlsx": {
        "prefix": "R", "file": "tests/playwright/cars/test_list.py"
    },
    "01_TaiLieu/TC_XoaThongTinXe.xlsx": {
        "prefix": "D", "file": "tests/playwright/cars/test_delete.py"
    }
}

CONTRACTS_EXCEL = "01_TaiLieu/TC_HopDong.xlsx"
CONTRACTS_SHEETS = {
    "3.1 Thêm HĐ": "tests/playwright/contracts/test_create.py",
    "3.2. Xem HĐ": "tests/playwright/contracts/test_search.py",
    "3.3. Sửa HĐ": "tests/playwright/contracts/test_edit.py",
    "3.4. Xóa HĐ": "tests/playwright/contracts/test_delete.py",
    "3.5. Gia hạn HĐ": "tests/playwright/contracts/test_extension.py",
    "3.6. Thêm GDTC": "tests/playwright/contracts/test_trans.py",
    "3.7. Trả xe": "tests/playwright/contracts/test_trans.py"
}

def extract_tc_data(sheet, car_prefix=None):
    rows = list(sheet.iter_rows(values_only=True))
    tc_data = {}
    
    header_idx = -1
    for i, row in enumerate(rows):
        row_str = [str(c) for c in row if c]
        if "ID" in row_str:
            header_idx = i
            break
    
    if header_idx == -1: return {}
    
    headers = [str(c).strip() if c else "" for c in rows[header_idx]]
    col_map = {
        "ID": -1, "Kịch bản thử nghiệm": -1, "Quy trình": -1, 
        "Kết quả mong đợi": -1, "Kết quả thực tế": -1, "Sub Module": -1
    }
    for i, h in enumerate(headers):
        if h in col_map: col_map[h] = i
    
    current_sub_module = ""

    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        
        # Track Sub Module for Cars GUI tests
        if col_map["Sub Module"] != -1 and row[col_map["Sub Module"]]:
            current_sub_module = str(row[col_map["Sub Module"]]).strip()

        raw_id = str(row[col_map["ID"]]) if col_map["ID"] != -1 and row[col_map["ID"]] is not None else ""
        summary = str(row[col_map["Kịch bản thử nghiệm"]]) if col_map["Kịch bản thử nghiệm"] != -1 and row[col_map["Kịch bản thử nghiệm"]] else ""
        
        if not raw_id and not summary: continue
        
        func_name = None
        # Mode 1: TC-ID style (Contracts)
        match = re.search(r"TC-([A-Z0-9-]+)", raw_id)
        if match:
            func_name = f"test_{match.group(0).replace('-', '_')}"
        # Mode 2: Numeric ID (Cars)
        elif raw_id.strip().isdigit() and car_prefix:
            num = int(raw_id.strip())
            prefix = car_prefix
            if "GUI" in current_sub_module:
                prefix = "G"
            func_name = f"test_TC_XE_{prefix}{num:02d}"

        if not func_name: continue
        
        steps = str(row[col_map["Quy trình"]]) if col_map["Quy trình"] != -1 and row[col_map["Quy trình"]] else ""
        expected = str(row[col_map["Kết quả mong đợi"]]) if col_map["Kết quả mong đợi"] != -1 and row[col_map["Kết quả mong đợi"]] else ""
        actual = str(row[col_map["Kết quả thực tế"]]) if col_map["Kết quả thực tế"] != -1 and row[col_map["Kết quả thực tế"]] else ""
        
        if func_name not in tc_data:
            tc_data[func_name] = {"summary": summary.strip(), "steps": steps.strip(), "expected": expected.strip(), "actual": actual.strip()}
        else:
            if steps: tc_data[func_name]["steps"] += "\n" + steps.strip()
            if expected: tc_data[func_name]["expected"] += "\n" + expected.strip()
            if actual: tc_data[func_name]["actual"] += "\n" + actual.strip()
            
    return tc_data

def update_file_docstrings(file_path, tc_data):
    if not os.path.exists(file_path): return
    with open(file_path, "r", encoding="utf-8") as f: content = f.read()
    
    updated_count = 0
    for func_name, data in tc_data.items():
        # Pattern to match: def func_name(...): followed by optional docstring
        pattern = r"(def\s+" + re.escape(func_name) + r"\s*\([^)]*\)\s*:\s*)(?:\n\s*\"\"\"(.*?)\"\"\"|(?=\n))"
        
        new_doc = f"\n    \"\"\"\n    Steps:\n    {data['summary']}\n    {data['steps']}\n    Expected:\n    {data['expected']}\n    Actual:\n    {data['actual']}\n    \"\"\""
        
        if re.search(pattern, content, re.DOTALL):
            content = re.sub(pattern, r"\1" + new_doc, content, flags=re.DOTALL)
            updated_count += 1

    with open(file_path, "w", encoding="utf-8") as f: f.write(content)
    print(f"Updated {file_path} ({updated_count} test cases)")

# --- PROCESS CONTRACTS ---
if os.path.exists(CONTRACTS_EXCEL):
    wb = openpyxl.load_workbook(CONTRACTS_EXCEL, data_only=True)
    for sheet_name, test_file in CONTRACTS_SHEETS.items():
        if sheet_name in wb.sheetnames:
            print(f"Processing Contracts sheet: {sheet_name}")
            data = extract_tc_data(wb[sheet_name])
            update_file_docstrings(test_file, data)

# --- PROCESS CARS ---
for car_excel, info in CARS_MAPPING.items():
    if os.path.exists(car_excel):
        print(f"Processing Cars Excel: {car_excel}")
        wb = openpyxl.load_workbook(car_excel, data_only=True)
        data = extract_tc_data(wb.active, car_prefix=info["prefix"])
        update_file_docstrings(info["file"], data)

print("Sync completed.")
